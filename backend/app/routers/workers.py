import io
import secrets
import string
from datetime import date, datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.core.database import get_db
from app.core.security import hash_password
from app.core.config import get_settings
from app.routers.deps import get_current_user, require_staff
from sqlalchemy import func as sa_func
from app.models.models import (
    AuditLog,
    DocumentChecklist,
    ChecklistStatus,
    Organisation,
    User,
    Worker,
    WorkerStage,
    UserRole,
)
from app.core.employment_status import employment_status_allowed
from app.core.org_option_lists import (
    effective_department_options,
    effective_onboarding_stage_options,
    effective_rtw_category_options,
    effective_work_location_options,
)
from app.schemas.schemas import (
    WorkerCreate,
    WorkerUpdate,
    WorkerOut,
    WorkerDetailOut,
    ProfilePhotoPresignOut,
)
from app.core.profile_photo_storage import (
    read_and_validate_upload,
    store_worker_profile_photo,
    delete_stored_object,
    presigned_get_url,
    load_profile_photo_bytes,
)
from app.routers.documents import create_checklist_for_worker

MANAGER_NO_PAYROLL = frozenset({UserRole.TENANT_ADMIN, UserRole.COMPLIANCE_MANAGER})

UK_RESIDENCE_COUNTRIES = frozenset({"england", "northern_ireland", "wales", "scotland", "outside_uk"})
SALARY_PAY_TYPES = frozenset({"hourly", "daily", "weekly", "monthly", "annual"})


def _age_years_from_dob(dob: date | datetime | None) -> int | None:
    if dob is None:
        return None
    d = dob.date() if isinstance(dob, datetime) else dob
    today = date.today()
    return today.year - d.year - ((today.month, today.day) < (d.month, d.day))


def _addr_snapshot(w: Worker) -> str:
    return "|".join(
        str(getattr(w, f, "") or "")
        for f in ("address_line_1", "address_line_2", "address_line_3", "postal_code", "uk_residence_country", "address")
    )


def _salary_snapshot(w: Worker) -> str:
    return f"{w.salary}|{w.salary_pay_type}"


def _audit_reporting_trigger(
    db: Session,
    current_user: User,
    worker_id: str,
    summary: str,
    before: str,
    after: str,
) -> None:
    db.add(
        AuditLog(
            user_id=current_user.id,
            user_email=current_user.email,
            user_role=current_user.role.value,
            action="UPDATE",
            entity_type="worker",
            entity_id=worker_id,
            details=f"{summary} (reporting trigger)",
            before_value=before[:8000],
            after_value=after[:8000],
        )
    )


def _validate_worker_fields_for_org(db: Session, org_id: str, data: dict) -> None:
    org = db.query(Organisation).filter(Organisation.id == org_id).first()
    if not org:
        return
    if "department" in data and data["department"] is not None:
        v = str(data["department"]).strip()
        if v and v not in set(effective_department_options(org)):
            raise HTTPException(status_code=400, detail="Invalid department for this organisation")
    if "work_location" in data and data["work_location"] is not None:
        v = str(data["work_location"]).strip()
        if v and v not in set(effective_work_location_options(org)):
            raise HTTPException(status_code=400, detail="Invalid work location for this organisation")
    if "hr_onboarding_stage" in data and data["hr_onboarding_stage"] is not None:
        v = str(data["hr_onboarding_stage"]).strip()
        if v and v not in set(effective_onboarding_stage_options(org)):
            raise HTTPException(status_code=400, detail="Invalid onboarding stage for this organisation")
    if "right_to_work_category" in data and data["right_to_work_category"] is not None:
        v = str(data["right_to_work_category"]).strip()
        if v and v not in set(effective_rtw_category_options(org)):
            raise HTTPException(status_code=400, detail="Invalid right to work category for this organisation")


def _sync_hr_stage_to_enum(worker: Worker) -> None:
    if not worker.hr_onboarding_stage:
        return
    s = worker.hr_onboarding_stage.lower()
    if "recruitment" in s:
        worker.stage = WorkerStage.RECRUITMENT
    elif "cos" in s and "assign" in s:
        worker.stage = WorkerStage.COS_ASSIGNMENT
    elif "pre" in s and "start" in s:
        worker.stage = WorkerStage.PRE_START
    elif "active" in s and "sponsor" in s:
        worker.stage = WorkerStage.ACTIVE_SPONSORSHIP
    elif "terminat" in s:
        worker.stage = WorkerStage.TERMINATED


def _maybe_redact_worker_list(worker: Worker, user: User) -> WorkerOut:
    w = WorkerOut.model_validate(worker)
    has_photo = bool(worker.profile_photo_s3_key or worker.profile_photo_data)
    if user.role in MANAGER_NO_PAYROLL:
        return w.model_copy(update={"salary": None, "has_profile_photo": has_photo})
    return w.model_copy(update={"has_profile_photo": has_photo})


def _maybe_redact_worker_detail(worker: Worker, user: User) -> WorkerDetailOut:
    out = WorkerDetailOut.model_validate(worker)
    has_photo = bool(worker.profile_photo_s3_key or worker.profile_photo_data)
    age = _age_years_from_dob(worker.date_of_birth)
    if user.role in MANAGER_NO_PAYROLL:
        return out.model_copy(
            update={
                "salary": None,
                "bank_account_number": None,
                "sort_code": None,
                "has_profile_photo": has_photo,
                "age_years": age,
            }
        )
    return out.model_copy(update={"has_profile_photo": has_photo, "age_years": age})

settings = get_settings()
DEFAULT_EMPLOYEE_PASSWORD = settings.DEFAULT_EMPLOYEE_PASSWORD


def _create_employee_user(
    db: Session,
    worker: Worker,
    organisation_id: str,
) -> User | None:
    """Auto-create a User account with EMPLOYEE role linked to a Worker.
    Returns the User, or None if the worker has no email.
    """
    if not worker.email:
        return None

    existing = db.query(User).filter(User.email == worker.email).first()
    if existing:
        if not existing.worker_id:
            existing.worker_id = worker.id
            existing.role = UserRole.EMPLOYEE
        return existing

    user = User(
        organisation_id=organisation_id,
        email=worker.email,
        hashed_password=hash_password(DEFAULT_EMPLOYEE_PASSWORD),
        full_name=worker.name,
        role=UserRole.EMPLOYEE,
        worker_id=worker.id,
        phone=worker.phone,
    )
    db.add(user)
    return user

router = APIRouter(prefix="/workers", tags=["workers"])

BULK_COLUMNS = [
    "first_name", "last_name", "job_title", "email", "phone", "nationality",
    "department", "soc_code", "salary", "route", "work_location",
    "start_date", "visa_expiry", "passport_expiry", "brp_expiry", "stage",
    "address", "postal_code", "date_of_birth", "place_of_birth", "country_of_birth",
    "gender", "ethnicity", "religion", "ni_number",
    "passport_number", "passport_place_of_issue", "passport_issue_date",
    "emergency_contact_name", "emergency_contact_phone",
    "next_of_kin_name", "next_of_kin_phone",
    "employee_id", "employee_type",
]

BULK_HEADERS = [
    "First Name *", "Last Name *", "Job Title *", "Email *", "Phone", "Nationality",
    "Department", "SOC Code", "Salary (£)", "Visa Route", "Work Location",
    "Start Date (YYYY-MM-DD)", "Visa Expiry (YYYY-MM-DD)",
    "Passport Expiry (YYYY-MM-DD)", "BRP Expiry (YYYY-MM-DD)", "Stage",
    "Home Address", "Postal Code", "Date of Birth (YYYY-MM-DD)",
    "Place of Birth", "Country of Birth",
    "Gender", "Ethnicity", "Religion", "NI Number",
    "Passport Number", "Passport Place of Issue", "Passport Issue Date (YYYY-MM-DD)",
    "Emergency Contact Name", "Emergency Contact Phone",
    "Next of Kin Name", "Next of Kin Phone",
    "Employee ID", "Employee Type",
]

SAMPLE_ROWS = [
    ["Claudia Manuel", "Canelhas Pinhao", "Live in Carer", "cpinhao@gmail.com", "7748519695",
     "Portugese", "Care", "6145", 25000, "Skilled Worker", "London",
     "2024-06-01", "2028-05-31", "2030-12-19", "2028-05-31", "active_sponsorship",
     "Rua dos Eucaliptos 98, Bairro da Encarnacao, LISBOA, Portugal", "1800-202",
     "1971-07-18", "Lisboa", "Portugal",
     "Female", "", "", "NJ645681B",
     "CH033822", "Portugal", "2025-12-19",
     "Maria Emilia Salves Canelhas Pinhao", "351936991149",
     "Maria Emilia Salves Canelhas Pinhao", "351936991149",
     "1", "migrant"],
    ["Tom", "Brown", "Data Scientist", "tom.brown@company.com", "+44 7700 900002",
     "Nigerian", "Analytics", "2425", 42000, "Skilled Worker", "Manchester Office",
     "2025-01-15", "2029-01-14", "2031-07-20", "2029-01-14", "recruitment",
     "12 Oxford Road, Manchester", "M1 5QA",
     "1990-03-25", "Lagos", "Nigeria",
     "Male", "", "", "CD789012E",
     "NG654321", "Lagos", "2023-03-25",
     "Funke Brown", "+44 7700 900003",
     "Funke Brown", "+44 7700 900003",
     "2", "migrant"],
]


# ── List workers ──

@router.get("", response_model=list[WorkerOut])
@router.get("/", response_model=list[WorkerOut], include_in_schema=False)
def list_workers(
    stage: str | None = None,
    status: str | None = None,
    search: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role == UserRole.EMPLOYEE:
        if current_user.worker_id:
            worker = db.query(Worker).filter(Worker.id == current_user.worker_id).first()
            return [worker] if worker else []
        return []

    query = db.query(Worker).filter(Worker.organisation_id == current_user.organisation_id)

    if stage:
        query = query.filter(Worker.stage == stage)
    if status:
        query = query.filter(Worker.status == status)
    if search:
        query = query.filter(Worker.name.ilike(f"%{search}%"))

    rows = query.order_by(Worker.created_at.desc()).all()
    return [_maybe_redact_worker_list(w, current_user) for w in rows]


# ── Compliance summary for all workers ──

@router.get("/compliance-summary")
def get_compliance_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_staff),
):
    rows = (
        db.query(
            DocumentChecklist.worker_id,
            DocumentChecklist.status,
            sa_func.count(DocumentChecklist.id),
        )
        .group_by(DocumentChecklist.worker_id, DocumentChecklist.status)
        .all()
    )

    summary: dict[str, dict] = {}
    for worker_id, st, cnt in rows:
        if worker_id not in summary:
            summary[worker_id] = {"total": 0, "verified": 0, "uploaded": 0, "rejected": 0}
        summary[worker_id]["total"] += cnt
        if st in (ChecklistStatus.VERIFIED, ChecklistStatus.NOT_APPLICABLE):
            summary[worker_id]["verified"] += cnt
        elif st == ChecklistStatus.UPLOADED:
            summary[worker_id]["uploaded"] += cnt
        elif st == ChecklistStatus.REJECTED:
            summary[worker_id]["rejected"] += cnt

    return summary


# ── Template download (must be before /{worker_id}) ──

@router.get("/template", response_class=StreamingResponse)
def download_template(current_user: User = Depends(require_staff)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B5DA8", end_color="2B5DA8", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D0D5DD"),
        right=Side(style="thin", color="D0D5DD"),
        top=Side(style="thin", color="D0D5DD"),
        bottom=Side(style="thin", color="D0D5DD"),
    )

    for col_idx, header in enumerate(BULK_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, row_data in enumerate(SAMPLE_ROWS, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    for col_idx in range(1, len(BULK_HEADERS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employee_bulk_template.xlsx"},
    )


# ── Bulk upload (must be before /{worker_id}) ──

@router.post("/bulk")
def bulk_upload(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_staff),
):
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx)")

    try:
        wb = load_workbook(file.file, read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read the Excel file. Make sure it is a valid .xlsx")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    if not rows:
        raise HTTPException(status_code=400, detail="The spreadsheet has no data rows")

    created = 0
    errors = []

    for row_num, row in enumerate(rows, 2):
        if not row or not row[0]:
            continue

        values = list(row) + [None] * (len(BULK_COLUMNS) - len(row))

        col = {k: values[i] for i, k in enumerate(BULK_COLUMNS)}

        first_name = str(col["first_name"]).strip() if col["first_name"] else ""
        last_name = str(col["last_name"]).strip() if col["last_name"] else ""
        job_title = str(col["job_title"]).strip() if col["job_title"] else ""
        email = str(col["email"]).strip().lower() if col["email"] else ""

        if (not first_name and not last_name) or not job_title:
            errors.append(f"Row {row_num}: first_name/last_name and job_title are required")
            continue
        if not email:
            errors.append(f"Row {row_num}: email is required")
            continue

        full_name = f"{first_name} {last_name}".strip()

        def parse_date(val):
            if val is None:
                return None
            if isinstance(val, datetime):
                return val.replace(tzinfo=timezone.utc) if val.tzinfo is None else val
            try:
                return datetime.strptime(str(val).strip(), "%Y-%m-%d").replace(tzinfo=timezone.utc)
            except ValueError:
                return None

        def opt_str(val):
            return str(val).strip() if val else None

        salary_raw = col["salary"]
        try:
            salary = float(salary_raw) if salary_raw else 0
        except (ValueError, TypeError):
            salary = 0

        worker = Worker(
            organisation_id=current_user.organisation_id,
            name=full_name,
            first_name=first_name or None,
            last_name=last_name or None,
            job_title=job_title,
            email=email,
            phone=opt_str(col["phone"]),
            nationality=opt_str(col["nationality"]),
            department=opt_str(col["department"]),
            soc_code=opt_str(col["soc_code"]),
            salary=salary,
            route=opt_str(col["route"]) or "Skilled Worker",
            work_location=opt_str(col["work_location"]),
            start_date=parse_date(col["start_date"]),
            visa_expiry=parse_date(col["visa_expiry"]),
            passport_expiry=parse_date(col["passport_expiry"]),
            brp_expiry=parse_date(col["brp_expiry"]),
            stage=opt_str(col["stage"]) or "recruitment",
            address=opt_str(col["address"]),
            postal_code=opt_str(col["postal_code"]),
            date_of_birth=parse_date(col["date_of_birth"]),
            place_of_birth=opt_str(col["place_of_birth"]),
            country_of_birth=opt_str(col["country_of_birth"]),
            gender=opt_str(col["gender"]),
            ethnicity=opt_str(col["ethnicity"]),
            religion=opt_str(col["religion"]),
            ni_number=opt_str(col["ni_number"]),
            passport_number=opt_str(col["passport_number"]),
            passport_place_of_issue=opt_str(col["passport_place_of_issue"]),
            passport_issue_date=parse_date(col["passport_issue_date"]),
            emergency_contact_name=opt_str(col["emergency_contact_name"]),
            emergency_contact_phone=opt_str(col["emergency_contact_phone"]),
            next_of_kin_name=opt_str(col["next_of_kin_name"]),
            next_of_kin_phone=opt_str(col["next_of_kin_phone"]),
            employee_id=opt_str(col["employee_id"]),
            employee_type=opt_str(col["employee_type"]),
        )
        db.add(worker)
        db.flush()
        create_checklist_for_worker(db, worker.id)
        _create_employee_user(db, worker, current_user.organisation_id)
        created += 1

    if created > 0:
        db.commit()

    return {
        "created": created,
        "errors": errors,
        "default_password": DEFAULT_EMPLOYEE_PASSWORD if created > 0 else None,
    }


# ── Profile photo (S3 or DB blob) — must be before /{worker_id} ──


def _worker_scope_or_403(
    db: Session,
    worker_id: str,
    current_user: User,
) -> Worker:
    worker = db.query(Worker).filter(
        Worker.id == worker_id,
        Worker.organisation_id == current_user.organisation_id,
    ).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    if current_user.role == UserRole.EMPLOYEE and current_user.worker_id != worker_id:
        raise HTTPException(status_code=403, detail="You can only view your own record")
    return worker


@router.get("/{worker_id}/profile-photo/presign", response_model=ProfilePhotoPresignOut)
def presign_profile_photo(
    worker_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Optional: presigned GET URL when using S3 (e.g. external viewers)."""
    worker = _worker_scope_or_403(db, worker_id, current_user)
    if not worker.profile_photo_s3_key:
        return ProfilePhotoPresignOut(url=None)
    url = presigned_get_url(worker.profile_photo_s3_key)
    return ProfilePhotoPresignOut(url=url)


@router.get("/{worker_id}/profile-photo")
def get_worker_profile_photo(
    worker_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Authenticated download of the profile image (works with S3 or DB blob)."""
    worker = _worker_scope_or_403(db, worker_id, current_user)
    loaded = load_profile_photo_bytes(worker)
    if not loaded:
        raise HTTPException(status_code=404, detail="No profile photo")
    body, mime = loaded
    return Response(
        content=body,
        media_type=mime,
        headers={"Cache-Control": "private, max-age=300"},
    )


@router.post("/{worker_id}/profile-photo", response_model=WorkerOut)
async def upload_worker_profile_photo(
    worker_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_staff),
):
    worker = db.query(Worker).filter(
        Worker.id == worker_id,
        Worker.organisation_id == current_user.organisation_id,
    ).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")

    raw, mime = await read_and_validate_upload(file)
    delete_stored_object(worker.profile_photo_s3_key)

    s3_key, blob = store_worker_profile_photo(
        organisation_id=current_user.organisation_id,
        worker_id=worker.id,
        file_bytes=raw,
        mime=mime,
    )
    worker.profile_photo_s3_key = s3_key
    worker.profile_photo_mime = mime
    worker.profile_photo_data = blob

    db.commit()
    db.refresh(worker)
    return worker


@router.delete("/{worker_id}/profile-photo", response_model=WorkerOut)
def delete_worker_profile_photo(
    worker_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_staff),
):
    worker = db.query(Worker).filter(
        Worker.id == worker_id,
        Worker.organisation_id == current_user.organisation_id,
    ).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    delete_stored_object(worker.profile_photo_s3_key)
    worker.profile_photo_s3_key = None
    worker.profile_photo_mime = None
    worker.profile_photo_data = None
    db.commit()
    db.refresh(worker)
    return worker


# ── Single worker CRUD (dynamic {worker_id} routes last) ──

@router.get("/{worker_id}", response_model=WorkerDetailOut)
def get_worker(
    worker_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role == UserRole.EMPLOYEE and current_user.worker_id != worker_id:
        raise HTTPException(status_code=403, detail="You can only view your own record")

    worker = db.query(Worker).filter(
        Worker.id == worker_id,
        Worker.organisation_id == current_user.organisation_id,
    ).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    return _maybe_redact_worker_detail(worker, current_user)


@router.post("", response_model=WorkerOut, status_code=status.HTTP_201_CREATED)
@router.post("/", response_model=WorkerOut, status_code=status.HTTP_201_CREATED, include_in_schema=False)
def create_worker(
    payload: WorkerCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_staff),
):
    data = payload.model_dump()
    if not data.get("name") and (data.get("first_name") or data.get("last_name")):
        data["name"] = f'{data.get("first_name", "") or ""} {data.get("last_name", "") or ""}'.strip()
    email = (data.get("email") or "").strip().lower() if data.get("email") else ""
    if not email:
        raise HTTPException(status_code=400, detail="Employee email is required")
    data["email"] = email
    es = (data.get("employment_status") or "Active").strip() or "Active"
    if not employment_status_allowed(db, current_user.organisation_id, es):
        raise HTTPException(status_code=400, detail="Invalid employment_status for this organisation")
    data["employment_status"] = es
    if data.get("uk_residence_country"):
        v = str(data["uk_residence_country"]).strip().lower()
        if v not in UK_RESIDENCE_COUNTRIES:
            raise HTTPException(status_code=400, detail="Invalid UK residence country")
        data["uk_residence_country"] = v
    if data.get("salary_pay_type"):
        v = str(data["salary_pay_type"]).strip().lower()
        if v not in SALARY_PAY_TYPES:
            raise HTTPException(status_code=400, detail="Invalid salary pay type")
        data["salary_pay_type"] = v
    else:
        data["salary_pay_type"] = "annual"
    _validate_worker_fields_for_org(db, current_user.organisation_id, data)
    if not data.get("hr_onboarding_stage") and data.get("stage"):
        st = str(data["stage"])
        data["hr_onboarding_stage"] = st.replace("_", " ").title()
    worker = Worker(organisation_id=current_user.organisation_id, **data)
    db.add(worker)
    db.flush()
    create_checklist_for_worker(db, worker.id)
    emp_user = _create_employee_user(db, worker, current_user.organisation_id)
    db.commit()
    db.refresh(worker)

    result = WorkerOut.model_validate(worker).model_dump()
    if emp_user:
        result["employee_login"] = {
            "email": emp_user.email,
            "default_password": DEFAULT_EMPLOYEE_PASSWORD,
        }
    return result


@router.patch("/{worker_id}", response_model=WorkerOut)
def update_worker(
    worker_id: str,
    payload: WorkerUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_staff),
):
    worker = db.query(Worker).filter(
        Worker.id == worker_id,
        Worker.organisation_id == current_user.organisation_id,
    ).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")

    pd = payload.model_dump(exclude_unset=True)
    if "uk_residence_country" in pd and pd["uk_residence_country"] is not None:
        v = str(pd["uk_residence_country"]).strip().lower()
        if v not in UK_RESIDENCE_COUNTRIES:
            raise HTTPException(status_code=400, detail="Invalid UK residence country")
        pd["uk_residence_country"] = v
    if "salary_pay_type" in pd and pd["salary_pay_type"] is not None:
        v = str(pd["salary_pay_type"]).strip().lower()
        if v not in SALARY_PAY_TYPES:
            raise HTTPException(status_code=400, detail="Invalid salary pay type")
        pd["salary_pay_type"] = v

    _validate_worker_fields_for_org(db, current_user.organisation_id, pd)

    before_addr = _addr_snapshot(worker)
    before_salary = _salary_snapshot(worker)

    for key, value in pd.items():
        if key == "employment_status":
            if value is None:
                continue
            v = str(value).strip()
            if not v:
                raise HTTPException(status_code=400, detail="employment_status cannot be empty")
            if not employment_status_allowed(db, current_user.organisation_id, v):
                raise HTTPException(status_code=400, detail="Invalid employment_status for this organisation")
            setattr(worker, "employment_status", v)
            continue
        setattr(worker, key, value)

    if "hr_onboarding_stage" in pd:
        _sync_hr_stage_to_enum(worker)

    after_addr = _addr_snapshot(worker)
    after_salary = _salary_snapshot(worker)
    if before_addr != after_addr:
        _audit_reporting_trigger(db, current_user, worker_id, "Address fields", before_addr, after_addr)
    if before_salary != after_salary:
        _audit_reporting_trigger(db, current_user, worker_id, "Salary / pay type", before_salary, after_salary)

    db.commit()
    db.refresh(worker)
    return worker


@router.delete("/{worker_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_worker(
    worker_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_staff),
):
    worker = db.query(Worker).filter(
        Worker.id == worker_id,
        Worker.organisation_id == current_user.organisation_id,
    ).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")

    db.delete(worker)
    db.commit()
